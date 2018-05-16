# -*- coding: utf-8 -*-

import logging, datetime,os,sys
from collections import OrderedDict
from openpyxl import load_workbook
from configs import config


# from config import PARMFILE_NAME, DEF_SHEET_NAME

#singleton mode via decorator 
def singleton(cls): 
	_instance = {}
	def _warper(*args, **kargs):
		if cls not in _instance:
			_instance[cls] = cls(*args, **kargs)
		return _instance[cls]
	return _warper

# Class definition

class RequestsLoader(object):
	"""docstring for ParameterLoader
	load the parameter file and retrun the list of request
	"""
	def __init__(self):
		self.parameter_filename = ''
		self.sheetname = ''
		self.wb = None
		self.sheet = None

	def load_workbook(self, parameter_filename=config.PARMFILE_NAME, sheetname=config.DEF_SHEET_NAME):

		self.parameter_filename = parameter_filename
		try:
			self.wb = load_workbook(self.parameter_filename)
		except Exception as e:
			logging.error('The error is %s, the dirname is %s' %(e, 
				os.path.abspath(self.parameter_filename)))

		try:
			self.sheet = self.wb[sheetname]
		except Exception as e:
			logging.error('the error is %s' %e)

	def get_records(self):

		if self.sheet is None:
			logging.error('Please init the loadrequest object')
			return False

		records = []
		for row in range(1, self.sheet.max_row+1):
			records.append(self.sheet[row])
		return records

	def get_requests(self):

		if self.sheet is None:
			logging.error('Please init the loadrequest object')
			return False

		records_list = []
		title_value = [col.value for col in self.sheet[2]]

		logging.info('The maxrow of parameter file is %s' %self.sheet.max_row )
		
		for row in range(3, self.sheet.max_row+1):
			if self.sheet[row][0].value:
				row_value = [ col.value for col in self.sheet[row]]
				records_list.append(OrderedDict(zip(title_value, row_value)))
		return records_list

	def get_requests_str(self):

		if self.sheet is None:
			logging.error('Please init the loadrequest object')
			return False

		records_list = []
		title_value = [col.value for col in self.sheet[2]]

		logging.debug('The maxrow of parameter file is %s' %self.sheet.max_row )
		
		for row in range(3, self.sheet.max_row+1):
			if self.sheet[row][0].value:
				row_value = []
				for col in self.sheet[row]:
					if isinstance(col.value, datetime.datetime):
						col.value = '-'.join([str(col.value.year), str(col.value.month), str(col.value.day)])
					row_value.append(col.value)
					
				records_list.append(OrderedDict(zip(title_value, row_value)))
		return records_list

	def get_email(self):

		if self.sheet is None:
			logging.error('Please init the loadrequest object')
			return False

		return self.sheet[1][1].value

if __name__ == '__main__':
	loadrequest = RequestsLoader()
	loadrequest.load_workbook('parameter_template_v1.xlsx')
	requests = loadrequest.get_requests_str()

	for request in requests:
		for key, value in request.items():
			print('%s : %s, type is %s' %(key, value, type(value)))

	email = loadrequest.get_email()

	print(email, type(email))